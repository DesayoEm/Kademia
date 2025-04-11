from fpdf import FPDF
import csv
import os
import openpyxl
from openpyxl.utils import get_column_letter


from uuid import UUID

from ...errors.export_errors import ExportFormatError
from ....config import config
from ....database.models import *
from .gather_data import GatherData
from ...errors.database_errors import EntityNotFoundError
from ...errors.staff_organisation_errors import RoleNotFoundError


class ExportService:
    def __init__(self, session):
        self.session = session
        self.gatherer = GatherData()
        self.export_dir = config.EXPORT_DIR

    def export_entity(
            self,
            entity_model,
            entity_id: UUID,
            export_format: str
    ) -> str:
        """
        Export an entity and its related data.

        Args:
            entity_model: Base SQLAlchemy model of the entity.
            entity_id (UUID): ID of the entity.
            export_format (str): Export format ("pdf", "csv", "excel").

        Returns:
            str: Path to the exported file.
        """
        error_map = {
            StaffRole: (RoleNotFoundError, "Role")
        }

        entity = self.session.get(entity_model, entity_id)

        if not entity:
            error_info = error_map.get(entity_model)

            if error_info:
                error_class, display_name = error_info
                raise error_class(identifier=entity_id, detail=f"{display_name} not found.")
            else:
                raise EntityNotFoundError(
                    entity_type=entity_model.__name__,
                    identifier=str(entity_id),
                    error="Object not found."
                )

        data, suffix = self.gatherer.gather(entity)

        if export_format == "pdf":
            return self.export_to_pdf(data, suffix)
        elif export_format == "csv":
            return self.export_to_csv(data, suffix)
        elif export_format == "excel":
            return self.export_to_excel(data, suffix)
        else:
            raise ExportFormatError(format_entry=export_format)



    def export_to_pdf(self, data: dict, suffix: str) -> str:
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=10)


        self.write_dict_to_pdf(pdf, data)

        os.makedirs(self.export_dir, exist_ok=True)

        filename = f"{self.export_dir}/{suffix}.pdf"
        pdf.output(filename)

        return filename


    def write_dict_to_pdf(self, pdf, data: dict, indent=0):
        """Recursively write a dictionary into the PDF."""
        for key, value in data.items():
            if isinstance(value, dict):
                pdf.cell(0, 10, f"{' ' * indent}{key}:", ln=True)
                self.write_dict_to_pdf(pdf, value, indent + 2)
            elif isinstance(value, list):
                pdf.cell(0, 10, f"{' ' * indent}{key}:", ln=True)
                for item in value:
                    if isinstance(item, dict):
                        self.write_dict_to_pdf(pdf, item, indent + 4)
                    else:
                        pdf.cell(0, 10, f"{' ' * (indent + 4)}- {item}", ln=True)
            else:
                pdf.cell(0, 10, f"{' ' * indent}{key}: {value}", ln=True)


    def export_to_csv(self, data: dict, suffix: str) -> str:
        os.makedirs(self.export_dir, exist_ok=True)

        filename = f"{self.export_dir}/{suffix}.csv"

        with open(filename, mode='w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            # Write the data
            self._write_dict_to_csv(writer, data)

        return filename

    def _write_dict_to_csv(self, writer, data: dict, parent_key=""):
        """
        Recursively write dictionary items to a CSV.
        """
        for key, value in data.items():
            full_key = f"{parent_key}.{key}" if parent_key else key

            if isinstance(value, dict):
                self._write_dict_to_csv(writer, value, parent_key=full_key)
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        self._write_dict_to_csv(writer, item, parent_key=full_key)
                    else:
                        writer.writerow([full_key, item])
            else:
                writer.writerow([full_key, value])

    def export_to_excel(self, data: dict, suffix: str) -> str:
        os.makedirs(self.export_dir, exist_ok=True)

        filename = f"{self.export_dir}/{suffix}.xlsx"  

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exported Data"

        self._write_dict_to_excel(ws, data)

        wb.save(filename)

        return filename

    def _write_dict_to_excel(self, ws, data: dict, row=1, col=1):
        """
        Recursively write dictionary items to an Excel worksheet.
        """
        for key, value in data.items():
            if isinstance(value, dict):
                ws.cell(row=row, column=col, value=key)
                row = self._write_dict_to_excel(ws, value, row + 1, col + 1)
            elif isinstance(value, list):
                ws.cell(row=row, column=col, value=key)
                for item in value:
                    if isinstance(item, dict):
                        row = self._write_dict_to_excel(ws, item, row + 1, col + 1)
                    else:
                        ws.cell(row=row, column=col + 1, value=item)
                        row += 1
            else:
                ws.cell(row=row, column=col, value=key)
                ws.cell(row=row, column=col + 1, value=value)
                row += 1

        return row