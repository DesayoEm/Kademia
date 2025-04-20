from fpdf import FPDF
from uuid import UUID
import csv
import os
import openpyxl

from V2.app.core.shared.errors.maps import error_map
from V2.app.config import config
from V2.app.core.shared.errors import EntityNotFoundError, ExportFormatError
from .gather_data import GatherData


class ExportService:
    """
    Service for exporting entities and their related data to various file formats.

    Provides methods to export database entities to PDF, CSV, or Excel formats,
    handling entity retrieval, data gathering, and file generation.

    Attributes:
        session: SQLAlchemy database session
        gatherer (GatherData): Service for gathering entity data for export
        export_dir (str): Directory path where exported files will be saved
    """

    def __init__(self, session):
        """
        Initialize the export service with a database session.

        Args:
            session: SQLAlchemy database session
        """
        self.session = session
        self.gatherer = GatherData()
        self.export_dir = config.EXPORT_DIR

    def export_entity(
            self, entity_model, entity_id: UUID, export_format: str) -> str:
        """
        Export an entity and its related data to the specified format.

        Retrieves the entity from the database, gathers its data using the
        GatherData service, and exports it to the requested format.

        Args:
            entity_model: SQLAlchemy model class for the entity
            entity_id (UUID): Unique identifier of the entity to export
            export_format (str): Format to export to ("pdf", "csv", "excel")

        Returns:
            str: Path to the exported file

        Raises:
            EntityNotFoundError: If the entity doesn't exist
            ExportFormatError: If the requested export format is not supported
        """
        entity = self.session.get(entity_model, entity_id)

        if not entity:
            error_details = error_map.get(entity_model)
            if error_details:
                _, display_name  = error_details

                raise EntityNotFoundError(
                    entity_model=entity_model,identifier=entity_id,
                    display_name=display_name, error = "Object not found during export."
                )
            else:
                raise EntityNotFoundError(
                    entity_model=entity_model, identifier=entity_id,
                    display_name="", error="Object not found during export."
                )


        data, file_name = self.gatherer.gather(entity)

        if export_format == "pdf":
            return self.export_to_pdf(data, file_name)
        elif export_format == "csv":
            return self.export_to_csv(data, file_name)
        elif export_format == "excel":
            return self.export_to_excel(data, file_name)
        else:
            raise ExportFormatError(format_entry=export_format)


    def export_to_pdf(self, data: dict, suffix: str) -> str:
        """
        Export data to a PDF file.

        Creates a PDF document containing the structured data from the entity.

        Args:
            data (dict): Structured data to include in the PDF
            suffix (str): Filename suffix for the exported file

        Returns:
            str: Path to the created PDF file
        """
        pdf = FPDF()
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()
        pdf.set_font("Arial", size=8)

        self.write_dict_to_pdf(pdf, data)

        os.makedirs(self.export_dir, exist_ok=True)

        filename = f"{self.export_dir}/{suffix}.pdf"
        pdf.output(filename)

        return filename


    def write_dict_to_pdf(self, pdf, data: dict, indent=0):
        """
        Recursively write a dictionary into the PDF.

        Formats dictionary keys and values with proper indentation and handles
        nested dictionaries and lists appropriately.

        Args:
            pdf: FPDF object to write to
            data (dict): Dictionary containing the data to write
            indent (int): Current indentation level for formatting
        """
        for key, value in data.items():
            if isinstance(value, dict):
                pdf.cell(0, 10, f"{' ' * indent}{key}:", ln=True)
                self.write_dict_to_pdf(pdf, value, indent + 2)
            elif isinstance(value, list):
                pdf.cell(0, 10, f"{' ' * indent}{key}:", ln=True)
                for item in value:
                    if isinstance(item, dict):
                        self.write_dict_to_pdf(pdf, item, indent + 4)
                    else:
                        pdf.cell(0, 10, f"{' ' * (indent + 4)}- {item}", ln=True)
            else:
                pdf.cell(0, 10, f"{' ' * indent}{key}: {value}", ln=True)


    def export_to_csv(self, data: dict, suffix: str) -> str:
        """
        Export data to a CSV file.

        Creates a CSV file containing the structured data from the entity.
        Handles nested structures by flattening keys with dot notation.

        Args:
            data (dict): Structured data to include in the CSV
            suffix (str): Filename suffix for the exported file

        Returns:
            str: Path to the created CSV file
        """
        os.makedirs(self.export_dir, exist_ok=True)

        filename = f"{self.export_dir}/{suffix}.csv"

        with open(filename, mode='w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)

            self._write_dict_to_csv(writer, data)

        return filename


    def _write_dict_to_csv(self, writer, data: dict, parent_key=""):
        """
        Recursively write dictionary items to a CSV.

        Flattens the dictionary hierarchy using dot notation for keys
        and writes each key-value pair as a row in the CSV.

        Args:
            writer: CSV writer object to write to
            data (dict): Dictionary containing the data to write
            parent_key (str): Parent key for nested dictionaries
        """
        for key, value in data.items():
            full_key = f"{parent_key}.{key}" if parent_key else key

            if isinstance(value, dict):
                self._write_dict_to_csv(writer, value, parent_key=full_key)
            elif isinstance(value, list):
                for item in value:
                    if isinstance(item, dict):
                        self._write_dict_to_csv(writer, item, parent_key=full_key)
                    else:
                        writer.writerow([full_key, item])
            else:
                writer.writerow([full_key, value])


    def export_to_excel(self, data: dict, suffix: str) -> str:
        """
        Export data to an Excel file.

        Creates an Excel workbook containing the structured data from the entity.

        Args:
            data (dict): Structured data to include in the Excel file
            suffix (str): Filename suffix for the exported file

        Returns:
            str: Path to the created Excel file
        """
        os.makedirs(self.export_dir, exist_ok=True)

        filename = f"{self.export_dir}/{suffix}.xlsx"

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exported Data"

        self._write_dict_to_excel(ws, data)

        wb.save(filename)

        return filename


    def _write_dict_to_excel(self, ws, data: dict, row=1, col=1):
        """
        Recursively write dictionary items to an Excel worksheet.

        Formats the data in a hierarchical structure in the worksheet,
        with proper indentation for nested items.

        Args:
            ws: Excel worksheet object to write to
            data (dict): Dictionary containing the data to write
            row (int): Current row position in the worksheet
            col (int): Current column position in the worksheet

        Returns:
            int: The next available row number after writing
        """
        for key, value in data.items():
            if isinstance(value, dict):
                ws.cell(row=row, column=col, value=key)
                row = self._write_dict_to_excel(ws, value, row + 1, col + 1)
            elif isinstance(value, list):
                ws.cell(row=row, column=col, value=key)
                for item in value:
                    if isinstance(item, dict):
                        row = self._write_dict_to_excel(ws, item, row + 1, col + 1)
                    else:
                        ws.cell(row=row, column=col + 1, value=item)
                        row += 1
            else:
                ws.cell(row=row, column=col, value=key)
                ws.cell(row=row, column=col + 1, value=value)
                row += 1

        return row